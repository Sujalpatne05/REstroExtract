"""
Exporter Engine Module

This module provides Excel export functionality for restaurant records.
It creates Excel files with data and metadata sheets, with proper formatting
and filename management.

Requirements: 5.1, 5.2, 5.3, 5.4, 6.1, 6.2, 6.3
"""

import os
from datetime import datetime
from typing import Dict, List, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.config import ConfigManager
from src.logger import LoggerSystem


class ExporterEngine:
    """
    Engine for exporting restaurant records to Excel format.
    
    Creates Excel files with data and metadata sheets, handles filename
    conflicts, and formats columns for readability.
    """
    
    def __init__(self, config: ConfigManager, logger: LoggerSystem):
        """
        Initialize the ExporterEngine.
        
        Args:
            config: Configuration manager instance
            logger: Logger system instance
        """
        self.config = config
        self.logger = logger
        self.output_directory = config.get("output_directory", "./output")
        
        # Ensure output directory exists
        os.makedirs(self.output_directory, exist_ok=True)
    
    def export_to_excel(
        self, records: List[Dict[str, Any]], stats: Dict[str, int]
    ) -> str:
        """
        Export records to Excel file with data and metadata sheets.
        
        Args:
            records: List of restaurant records to export
            stats: Dictionary with extraction statistics
            
        Returns:
            Path to the created Excel file
            
        Requirements: 5.1, 5.2, 5.3, 5.4, 6.1, 6.2, 6.3
        """
        try:
            # Check if append mode is enabled
            append_mode = self.config.get("append_mode", False)
            
            if append_mode:
                return self._append_to_excel(records, stats)
            else:
                return self._create_new_excel(records, stats)
        
        except Exception as e:
            self.logger.error(
                "Failed to export to Excel",
                exception=e,
                context={"records": len(records)}
            )
            raise
    
    def _create_new_excel(
        self, records: List[Dict[str, Any]], stats: Dict[str, int]
    ) -> str:
        """
        Create a new Excel file with data and metadata sheets.
        
        Args:
            records: List of restaurant records to export
            stats: Dictionary with extraction statistics
            
        Returns:
            Path to the created Excel file
        """
        # Create workbook
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove default sheet
        
        # Create data sheet
        self.create_data_sheet(workbook, records)
        
        # Create metadata sheet
        self.create_metadata_sheet(workbook, stats)
        
        # Generate filename
        state = self.config.get("target_state", "Maharashtra")
        filename = self.generate_filename(state)
        filepath = os.path.join(self.output_directory, filename)
        
        # Handle filename conflicts
        filepath = self._handle_filename_conflict(filepath)
        
        # Save workbook
        workbook.save(filepath)
        
        self.logger.info(
            f"Excel file exported successfully",
            context={"filepath": filepath, "records": len(records)}
        )
        
        return filepath
    
    def _append_to_excel(
        self, records: List[Dict[str, Any]], stats: Dict[str, int]
    ) -> str:
        """
        Append records to an existing Excel file or create new one if it doesn't exist.
        
        Args:
            records: List of restaurant records to append
            stats: Dictionary with extraction statistics
            
        Returns:
            Path to the Excel file
        """
        from openpyxl import load_workbook
        
        state = self.config.get("target_state", "Maharashtra")
        
        # Generate filename without timestamp for consistent file
        timestamp = datetime.now().strftime("%Y%m%d")
        filename = f"FSSAI_Restaurants_{state}.xlsx"
        filepath = os.path.join(self.output_directory, filename)
        
        # Check if file exists
        if os.path.exists(filepath):
            # Load existing workbook
            workbook = load_workbook(filepath)
            sheet = workbook["Data"]
            
            # Get the last row number
            last_row = sheet.max_row
            
            # Define field names
            field_names = [
                "business_name",
                "license_number",
                "license_type",
                "business_type",
                "district",
                "city_town",
                "pin_code",
                "issue_date",
                "valid_till",
                "owner_contact",
                "mobile",
                "email",
                "address",
                "state",
            ]
            
            # Append new records
            for row_idx, record in enumerate(records, start=last_row + 1):
                for col_idx, field_name in enumerate(field_names, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.value = record.get(field_name, "")
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            # Update metadata sheet
            meta_sheet = workbook["Metadata"]
            
            # Update total records
            total_records = sheet.max_row - 1  # Exclude header
            meta_sheet.cell(row=1, column=2).value = datetime.now().isoformat()
            meta_sheet.cell(row=2, column=2).value = total_records
            meta_sheet.cell(row=3, column=2).value = stats.get("total_validated", 0)
            meta_sheet.cell(row=4, column=2).value = stats.get("total_rejected", 0)
            meta_sheet.cell(row=5, column=2).value = stats.get("total_duplicates_removed", 0)
            
            self.logger.info(
                f"Records appended to existing Excel file",
                context={"filepath": filepath, "new_records": len(records), "total_records": total_records}
            )
        else:
            # Create new workbook
            workbook = Workbook()
            workbook.remove(workbook.active)
            
            # Create data sheet
            self.create_data_sheet(workbook, records)
            
            # Create metadata sheet
            self.create_metadata_sheet(workbook, stats)
            
            self.logger.info(
                f"New Excel file created in append mode",
                context={"filepath": filepath, "records": len(records)}
            )
        
        # Save workbook
        workbook.save(filepath)
        
        return filepath
    
    def create_data_sheet(
        self, workbook: Workbook, records: List[Dict[str, Any]]
    ) -> None:
        """
        Create data sheet with headers and records.
        
        Args:
            workbook: Openpyxl workbook instance
            records: List of restaurant records
            
        Requirements: 5.1, 5.2
        """
        sheet = workbook.create_sheet("Data")
        
        # Define headers (14 fields)
        headers = [
            "Business Name",
            "License Number",
            "License Type",
            "Business Type",
            "District",
            "City/Town",
            "Pin Code",
            "Issue Date",
            "Valid Till",
            "Owner/Contact",
            "Mobile",
            "Email",
            "Address",
            "State",
        ]
        
        # Add headers to first row
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            # Format header row
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Map record fields to columns
        field_names = [
            "business_name",
            "license_number",
            "license_type",
            "business_type",
            "district",
            "city_town",
            "pin_code",
            "issue_date",
            "valid_till",
            "owner_contact",
            "mobile",
            "email",
            "address",
            "state",
        ]
        
        # Add records to sheet
        for row_idx, record in enumerate(records, start=2):
            for col_idx, field_name in enumerate(field_names, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = record.get(field_name, "")
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # Format columns
        self.format_columns(sheet)
    
    def create_metadata_sheet(
        self, workbook: Workbook, stats: Dict[str, int]
    ) -> None:
        """
        Create metadata sheet with extraction statistics.
        
        Args:
            workbook: Openpyxl workbook instance
            stats: Dictionary with extraction statistics
            
        Requirements: 5.4
        """
        sheet = workbook.create_sheet("Metadata")
        
        # Metadata entries
        metadata = [
            ("Generation Date", datetime.now().isoformat()),
            ("Total Records Extracted", stats.get("total_extracted", 0)),
            ("Total Records Validated", stats.get("total_validated", 0)),
            ("Total Records Rejected", stats.get("total_rejected", 0)),
            ("Duplicates Removed", stats.get("total_duplicates_removed", 0)),
            ("State Filter Applied", self.config.get("target_state", "Maharashtra")),
            ("Extraction Status", stats.get("extraction_status", "Unknown")),
        ]
        
        # Add metadata to sheet
        for row_idx, (key, value) in enumerate(metadata, start=1):
            # Key column
            key_cell = sheet.cell(row=row_idx, column=1)
            key_cell.value = key
            key_cell.font = Font(bold=True)
            key_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            # Value column
            value_cell = sheet.cell(row=row_idx, column=2)
            value_cell.value = value
            value_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Set column widths
        sheet.column_dimensions["A"].width = 25
        sheet.column_dimensions["B"].width = 40
    
    def format_columns(self, sheet) -> None:
        """
        Format columns for readability.
        
        Args:
            sheet: Openpyxl worksheet instance
            
        Requirements: 5.3
        """
        # Define column widths
        column_widths = {
            1: 25,   # Business Name
            2: 18,   # License Number
            3: 18,   # License Type
            4: 18,   # Business Type
            5: 15,   # District
            6: 15,   # City/Town
            7: 12,   # Pin Code
            8: 15,   # Issue Date
            9: 15,   # Valid Till
            10: 20,  # Owner/Contact
            11: 15,  # Mobile
            12: 25,  # Email
            13: 30,  # Address
            14: 15,  # State
        }
        
        # Apply column widths
        for col_idx, width in column_widths.items():
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Set row height for header
        sheet.row_dimensions[1].height = 30
    
    def generate_filename(self, state: str) -> str:
        """
        Generate filename with pattern FSSAI_Restaurants_[STATE]_[YYYYMMDD_HHMMSS].xlsx
        
        Args:
            state: State name for the filename
            
        Returns:
            Generated filename
            
        Requirements: 6.1
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"FSSAI_Restaurants_{state}_{timestamp}.xlsx"
        return filename
    
    def _handle_filename_conflict(self, filepath: str) -> str:
        """
        Handle filename conflicts by appending sequence numbers.
        
        If a file already exists at the given path, appends a sequence number
        to create a unique filename.
        
        Args:
            filepath: Original file path
            
        Returns:
            Unique file path (original or with sequence number appended)
            
        Requirements: 6.3
        """
        if not os.path.exists(filepath):
            return filepath
        
        # File exists, append sequence number
        base, ext = os.path.splitext(filepath)
        sequence = 1
        
        while True:
            new_filepath = f"{base}_{sequence}{ext}"
            if not os.path.exists(new_filepath):
                return new_filepath
            sequence += 1
