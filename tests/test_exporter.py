"""
Unit tests for ExporterEngine module.

Tests Excel export functionality, filename generation, and metadata creation.
"""

import os
import pytest
import tempfile
from datetime import datetime
from openpyxl import load_workbook

from src.config import ConfigManager
from src.logger import LoggerSystem
from src.exporter import ExporterEngine


class TestExporterEngine:
    """Test suite for ExporterEngine."""
    
    @pytest.fixture
    def temp_dir(self):
        """Create a temporary directory for test files."""
        with tempfile.TemporaryDirectory() as tmpdir:
            yield tmpdir
    
    @pytest.fixture
    def config(self, temp_dir):
        """Create a ConfigManager instance with temp directory."""
        config = ConfigManager()
        config.config = {
            "target_state": "Maharashtra",
            "output_directory": temp_dir,
            "request_timeout": 30,
            "rate_limit_delay": 1.0,
            "max_retries": 3,
            "log_file": os.path.join(temp_dir, "test.log"),
            "log_level": "INFO",
        }
        return config
    
    @pytest.fixture
    def logger(self, temp_dir):
        """Create a LoggerSystem instance."""
        logger = LoggerSystem(
            log_file=os.path.join(temp_dir, "test.log"),
            log_level="INFO"
        )
        yield logger
        # Close the logger to release file handles
        if hasattr(logger, 'logger') and logger.logger:
            for handler in logger.logger.handlers[:]:
                handler.close()
                logger.logger.removeHandler(handler)
    
    @pytest.fixture
    def exporter(self, config, logger):
        """Create an ExporterEngine instance."""
        return ExporterEngine(config, logger)
    
    def test_exporter_initialization(self, exporter, temp_dir):
        """Test ExporterEngine initialization."""
        assert exporter.output_directory == temp_dir
        assert os.path.exists(temp_dir)
    
    def test_generate_filename_format(self, exporter):
        """Test filename generation with correct format."""
        filename = exporter.generate_filename("Maharashtra")
        
        # Check format: FSSAI_Restaurants_[STATE]_[YYYYMMDD_HHMMSS].xlsx
        assert filename.startswith("FSSAI_Restaurants_Maharashtra_")
        assert filename.endswith(".xlsx")
        
        # Extract timestamp part
        timestamp_part = filename.replace("FSSAI_Restaurants_Maharashtra_", "").replace(".xlsx", "")
        
        # Verify timestamp format (YYYYMMDD_HHMMSS)
        assert len(timestamp_part) == 15  # 8 digits + _ + 6 digits
        assert timestamp_part[8] == "_"
    
    def test_generate_filename_different_states(self, exporter):
        """Test filename generation with different states."""
        filename1 = exporter.generate_filename("Maharashtra")
        filename2 = exporter.generate_filename("Karnataka")
        
        assert "Maharashtra" in filename1
        assert "Karnataka" in filename2
        assert filename1 != filename2
    
    def test_handle_filename_conflict_no_conflict(self, exporter, temp_dir):
        """Test filename conflict handling when no conflict exists."""
        filepath = os.path.join(temp_dir, "test.xlsx")
        result = exporter._handle_filename_conflict(filepath)
        
        assert result == filepath
    
    def test_handle_filename_conflict_with_conflict(self, exporter, temp_dir):
        """Test filename conflict handling when file exists."""
        filepath = os.path.join(temp_dir, "test.xlsx")
        
        # Create the file
        with open(filepath, "w") as f:
            f.write("test")
        
        result = exporter._handle_filename_conflict(filepath)
        
        # Should append sequence number
        assert result != filepath
        assert "test_1.xlsx" in result
    
    def test_handle_filename_conflict_multiple(self, exporter, temp_dir):
        """Test filename conflict handling with multiple conflicts."""
        filepath = os.path.join(temp_dir, "test.xlsx")
        
        # Create multiple files
        with open(filepath, "w") as f:
            f.write("test")
        with open(os.path.join(temp_dir, "test_1.xlsx"), "w") as f:
            f.write("test")
        
        result = exporter._handle_filename_conflict(filepath)
        
        # Should append sequence number 2
        assert "test_2.xlsx" in result
    
    def test_export_to_excel_creates_file(self, exporter, temp_dir):
        """Test that export_to_excel creates a file."""
        records = [
            {
                "business_name": "Restaurant A",
                "license_number": "LIC001",
                "license_type": "State",
                "business_type": "Restaurant",
                "district": "Mumbai",
                "city_town": "Mumbai",
                "pin_code": "400001",
                "issue_date": "01-01-2023",
                "valid_till": "31-12-2024",
                "owner_contact": "John Doe",
                "mobile": "9999999999",
                "email": "john@example.com",
                "address": "123 Main St",
                "state": "Maharashtra",
            }
        ]
        stats = {
            "total_extracted": 1,
            "total_validated": 1,
            "total_rejected": 0,
            "total_duplicates_removed": 0,
            "extraction_status": "Success",
        }
        
        filepath = exporter.export_to_excel(records, stats)
        
        assert os.path.exists(filepath)
        assert filepath.endswith(".xlsx")
    
    def test_export_to_excel_data_sheet_structure(self, exporter, temp_dir):
        """Test that data sheet has correct structure."""
        records = [
            {
                "business_name": "Restaurant A",
                "license_number": "LIC001",
                "license_type": "State",
                "business_type": "Restaurant",
                "district": "Mumbai",
                "city_town": "Mumbai",
                "pin_code": "400001",
                "issue_date": "01-01-2023",
                "valid_till": "31-12-2024",
                "owner_contact": "John Doe",
                "mobile": "9999999999",
                "email": "john@example.com",
                "address": "123 Main St",
                "state": "Maharashtra",
            }
        ]
        stats = {
            "total_extracted": 1,
            "total_validated": 1,
            "total_rejected": 0,
            "total_duplicates_removed": 0,
            "extraction_status": "Success",
        }
        
        filepath = exporter.export_to_excel(records, stats)
        
        # Load and verify workbook
        workbook = load_workbook(filepath)
        assert "Data" in workbook.sheetnames
        
        data_sheet = workbook["Data"]
        
        # Check headers
        headers = [cell.value for cell in data_sheet[1]]
        expected_headers = [
            "Business Name",
            "License Number",
            "License Type",
            "Business Type",
            "District",
            "City/Town",
            "Pin Code",
            "Issue Date",
            "Valid Till",
            "Owner/Contact",
            "Mobile",
            "Email",
            "Address",
            "State",
        ]
        assert headers == expected_headers
        
        # Check data row
        data_row = [cell.value for cell in data_sheet[2]]
        assert data_row[0] == "Restaurant A"
        assert data_row[1] == "LIC001"
        assert data_row[13] == "Maharashtra"
    
    def test_export_to_excel_metadata_sheet(self, exporter, temp_dir):
        """Test that metadata sheet contains correct information."""
        records = [
            {
                "business_name": "Restaurant A",
                "license_number": "LIC001",
                "license_type": "State",
                "business_type": "Restaurant",
                "district": "Mumbai",
                "city_town": "Mumbai",
                "pin_code": "400001",
                "issue_date": "01-01-2023",
                "valid_till": "31-12-2024",
                "owner_contact": "John Doe",
                "mobile": "9999999999",
                "email": "john@example.com",
                "address": "123 Main St",
                "state": "Maharashtra",
            }
        ]
        stats = {
            "total_extracted": 1,
            "total_validated": 1,
            "total_rejected": 0,
            "total_duplicates_removed": 0,
            "extraction_status": "Success",
        }
        
        filepath = exporter.export_to_excel(records, stats)
        
        # Load and verify workbook
        workbook = load_workbook(filepath)
        assert "Metadata" in workbook.sheetnames
        
        metadata_sheet = workbook["Metadata"]
        
        # Check metadata entries
        metadata_dict = {}
        for row in metadata_sheet.iter_rows(min_row=1, values_only=True):
            if row[0] and row[1] is not None:
                metadata_dict[row[0]] = row[1]
        
        assert metadata_dict["Total Records Extracted"] == 1
        assert metadata_dict["Total Records Validated"] == 1
        assert metadata_dict["Total Records Rejected"] == 0
        assert metadata_dict["Duplicates Removed"] == 0
        assert metadata_dict["State Filter Applied"] == "Maharashtra"
        assert metadata_dict["Extraction Status"] == "Success"
    
    def test_export_to_excel_multiple_records(self, exporter, temp_dir):
        """Test export with multiple records."""
        records = [
            {
                "business_name": f"Restaurant {i}",
                "license_number": f"LIC{i:03d}",
                "license_type": "State",
                "business_type": "Restaurant",
                "district": "Mumbai",
                "city_town": "Mumbai",
                "pin_code": "400001",
                "issue_date": "01-01-2023",
                "valid_till": "31-12-2024",
                "owner_contact": f"Owner {i}",
                "mobile": "9999999999",
                "email": f"owner{i}@example.com",
                "address": "123 Main St",
                "state": "Maharashtra",
            }
            for i in range(1, 6)
        ]
        stats = {
            "total_extracted": 5,
            "total_validated": 5,
            "total_rejected": 0,
            "total_duplicates_removed": 0,
            "extraction_status": "Success",
        }
        
        filepath = exporter.export_to_excel(records, stats)
        
        # Load and verify workbook
        workbook = load_workbook(filepath)
        data_sheet = workbook["Data"]
        
        # Check number of rows (header + 5 records)
        assert data_sheet.max_row == 6
    
    def test_export_to_excel_empty_records(self, exporter, temp_dir):
        """Test export with empty records list."""
        records = []
        stats = {
            "total_extracted": 0,
            "total_validated": 0,
            "total_rejected": 0,
            "total_duplicates_removed": 0,
            "extraction_status": "Success",
        }
        
        filepath = exporter.export_to_excel(records, stats)
        
        # Load and verify workbook
        workbook = load_workbook(filepath)
        data_sheet = workbook["Data"]
        
        # Should have only header row
        assert data_sheet.max_row == 1
    
    def test_export_to_excel_missing_fields(self, exporter, temp_dir):
        """Test export with records missing some fields."""
        records = [
            {
                "business_name": "Restaurant A",
                "license_number": "LIC001",
                # Missing other fields
            }
        ]
        stats = {
            "total_extracted": 1,
            "total_validated": 1,
            "total_rejected": 0,
            "total_duplicates_removed": 0,
            "extraction_status": "Success",
        }
        
        filepath = exporter.export_to_excel(records, stats)
        
        # Load and verify workbook
        workbook = load_workbook(filepath)
        data_sheet = workbook["Data"]
        
        # Should have header + 1 data row
        assert data_sheet.max_row == 2
        
        # Check that missing fields are empty
        data_row = [cell.value for cell in data_sheet[2]]
        assert data_row[0] == "Restaurant A"
        assert data_row[1] == "LIC001"
        assert data_row[2] is None  # Missing license_type
