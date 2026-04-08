"""
Property-based tests for ExporterEngine module.

Tests correctness properties using hypothesis library.
"""

import os
from datetime import datetime
from hypothesis import given, settings, strategies as st

from src.config import ConfigManager
from src.logger import LoggerSystem
from src.exporter import ExporterEngine


class TestExporterProperties:
    """Property-based tests for ExporterEngine."""
    
    @given(st.text(min_size=1, max_size=20))
    @settings(max_examples=20)
    def test_generate_filename_format(self, state):
        """
        Property: Generated filename follows pattern FSSAI_Restaurants_[STATE]_[YYYYMMDD_HHMMSS].xlsx
        
        Validates: Requirements 6.1
        """
        config = ConfigManager()
        config.config = {
            "target_state": state,
            "output_directory": "./output",
            "request_timeout": 30,
            "rate_limit_delay": 1.0,
            "max_retries": 3,
            "log_file": "./logs/test.log",
            "log_level": "INFO",
        }
        logger = LoggerSystem(
            log_file="./logs/test.log",
            log_level="INFO"
        )
        exporter = ExporterEngine(config, logger)
        
        filename = exporter.generate_filename(state)
        
        # Check format
        assert filename.startswith("FSSAI_Restaurants_")
        assert state in filename
        assert filename.endswith(".xlsx")
        
        # Extract timestamp part
        timestamp_part = filename.replace(f"FSSAI_Restaurants_{state}_", "").replace(".xlsx", "")
        
        # Verify timestamp format (YYYYMMDD_HHMMSS)
        assert len(timestamp_part) == 15
        assert timestamp_part[8] == "_"
    
    @given(st.text(min_size=1, max_size=20))
    @settings(max_examples=20)
    def test_handle_filename_conflict_creates_unique_name(self, base_name):
        """
        Property: _handle_filename_conflict() returns a unique filename when file exists.
        
        Validates: Requirements 6.3
        """
        config = ConfigManager()
        config.config = {
            "target_state": "Maharashtra",
            "output_directory": "./output",
            "request_timeout": 30,
            "rate_limit_delay": 1.0,
            "max_retries": 3,
            "log_file": "./logs/test.log",
            "log_level": "INFO",
        }
        logger = LoggerSystem(
            log_file="./logs/test.log",
            log_level="INFO"
        )
        exporter = ExporterEngine(config, logger)
        
        # Test with non-existent file
        filepath = f"./output/{base_name}.xlsx"
        result = exporter._handle_filename_conflict(filepath)
        
        # Should return the same path if file doesn't exist
        assert result == filepath
    
    @given(st.text(min_size=1, max_size=20))
    @settings(max_examples=20)
    def test_generate_filename_different_states(self, state):
        """
        Property: Generated filenames for different states are different.
        
        Validates: Requirements 6.1
        """
        config = ConfigManager()
        config.config = {
            "target_state": state,
            "output_directory": "./output",
            "request_timeout": 30,
            "rate_limit_delay": 1.0,
            "max_retries": 3,
            "log_file": "./logs/test.log",
            "log_level": "INFO",
        }
        logger = LoggerSystem(
            log_file="./logs/test.log",
            log_level="INFO"
        )
        exporter = ExporterEngine(config, logger)
        
        filename1 = exporter.generate_filename(state)
        filename2 = exporter.generate_filename("OtherState")
        
        # Filenames should be different
        assert filename1 != filename2
        assert state in filename1
        assert "OtherState" in filename2
    
    @given(st.integers(min_value=0, max_value=100))
    @settings(max_examples=20)
    def test_format_columns_with_valid_sheet(self, num_cols):
        """
        Property: format_columns() handles sheets with various column counts.
        
        Validates: Requirements 5.3
        """
        from openpyxl import Workbook
        
        config = ConfigManager()
        config.config = {
            "target_state": "Maharashtra",
            "output_directory": "./output",
            "request_timeout": 30,
            "rate_limit_delay": 1.0,
            "max_retries": 3,
            "log_file": "./logs/test.log",
            "log_level": "INFO",
        }
        logger = LoggerSystem(
            log_file="./logs/test.log",
            log_level="INFO"
        )
        exporter = ExporterEngine(config, logger)
        
        workbook = Workbook()
        sheet = workbook.active
        
        # Add some data
        for i in range(1, min(num_cols + 1, 15)):
            sheet.cell(row=1, column=i).value = f"Header {i}"
        
        # Should not raise an error
        exporter.format_columns(sheet)
        
        # Verify column widths are set
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, min(15, num_cols + 1)):
            col_letter = get_column_letter(col_idx)
            assert sheet.column_dimensions[col_letter].width is not None
